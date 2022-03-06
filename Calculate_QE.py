#!/usr/bin/env python
# coding: utf-8

# In[1]:


import torch
import openpyxl
import numpy as np 
from tqdm import tqdm
from tqe import TQE


# In[ ]:

# TQE 측정
wb = openpyxl.load_workbook('TQE(5-Epoch).xlsx')
sheet = wb['model_8']
# source : 번역 원문
# target : 기계 번역 문장

for i in range(500, 1002): 
    target = [] 
    source = []
    candidate = sheet.cell(row = i, column = 11).value.capitalize()
    reference = sheet.cell(row = i, column = 3).value
    print(i, ':', candidate)

    # Translation Quality Estimator (QE)
    # https://github.com/theRay07/Translation-Quality-Estimator
    target.append(candidate)
    source.append(reference)
    model = TQE('LaBSE')
    cos_sim_values = model.fit(source, target)
    sheet.cell(row = i, column = 12).value = cos_sim_values[0]
    
wb.save('TQE(5-Epoch).xlsx')


#################################################################################

# lang_1 = ["my son s name is tom ."]
# lang_2 = ["El nombre de mi hijo es Tom."]

# model = TQE('LaBSE')
# cos_sim_values = model.fit(lang_1, lang_2)
# print(cos_sim_values)

#################################################################################

# In[ ]:

# TQE 평균 값
# wb = openpyxl.load_workbook('TQE(5-Epoch).xlsx')
# sheet = wb['model_15']
# qe = []
# qe2 = []
# qe3 = []
# qe4 = []


# for i in range(2, 1002):
#     qe.append(float(sheet.cell(row = i, column = 8).value))
#     qe2.append(float(sheet.cell(row = i, column = 12).value))
#     qe3.append(float(sheet.cell(row = i, column = 16).value))
#     qe4.append(float(sheet.cell(row = i, column = 20).value))

    
# qe = np.array(qe)
# qe2 = np.array(qe2)
# qe3 = np.array(qe3)
# qe4 = np.array(qe4)

# sheet.cell(row = 1002, column = 8).value = qe.mean()
# sheet.cell(row = 1002, column = 12).value = qe2.mean()
# sheet.cell(row = 1002, column = 16).value = qe3.mean()
# sheet.cell(row = 1002, column = 20).value = qe4.mean()


# wb.save('TQE(5-Epoch).xlsx')


# no word 제거
# wb = openpyxl.load_workbook('TQE(5-Epoch).xlsx')
# sheet = wb['model_2']

# for i in range(2, 1002):
#     candidate = sheet.cell(row = i, column = 7).value
#     if candidate == ' ':
#         sheet.cell(row = i, column = 8).value = 0
#         sheet.cell(row = i, column = 12).value = 0
#         sheet.cell(row = i, column = 16).value = 0
#         sheet.cell(row = i, column = 20).value = 0


# wb.save('TQE(5-Epoch).xlsx')